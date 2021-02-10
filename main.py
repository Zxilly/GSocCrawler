from concurrent.futures.thread import ThreadPoolExecutor

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from tqdm import tqdm

from func import toStr

proxies = {
    'http': 'http://localhost:11000',
    'https': 'http://localhost:11000',
}

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/88.0.4324.96 Safari/537.36 Edg/88.0.705.56'}

proxySession = requests.session()
proxySession.proxies = proxies
proxySession.headers = headers

baseURL = 'https://summerofcode.withgoogle.com'
targetURL = baseURL + "/archive/2020/organizations/"

GSocMainPage = proxySession.get(targetURL, timeout=3).content

GSocMainBS = BeautifulSoup(GSocMainPage, 'lxml')

GSocOrganizationsBS = GSocMainBS.find(class_='organization-list-container')

GSocOrganizationNumber = len(GSocOrganizationsBS.find_all('li'))

print("Found %d organization." % GSocOrganizationNumber)

GSocOrganizationsInfoList = []
CrawlerErrorList = []
AllTechnology = set()


def infoGet(oneOrganizationBS):
    oneOrganizationURL = baseURL + oneOrganizationBS.a['href']
    oneOrganizationPage = proxySession.get(oneOrganizationURL).content
    oneOrganizationBS = BeautifulSoup(oneOrganizationPage, 'lxml')

    title = str(oneOrganizationBS.find(class_='banner__title').string)

    # print(oneOrganizationBS.prettify())
    metasBS = oneOrganizationBS.find(class_='org__meta')
    # print(metasBS.prettify())
    if metasBS is None:
        # print('%s meet error, continue' % title)
        CrawlerErrorList.append({
            'title': title,
            'url': oneOrganizationURL
        })
        return
    idealistURL = metasBS.find(class_='org__button-container').find('md-button')['href']
    technologyList = []
    for oneTechnologyBS in metasBS.find_all(class_='organization__tag--technology'):
        technology = str(oneTechnologyBS.string).strip()
        AllTechnology.add(technology)
        technologyList.append(technology)
    category = str(metasBS.find(class_='organization__tag--category').a.string)
    topicsList = []
    for oneTopicBS in metasBS.find_all(class_='organization__tag--topic'):
        topicsList.append(str(oneTopicBS.string).strip())

    GSocOrganizationsInfoList.append({
        'title': title,
        # 'url': '=HYPERLINK("{}", "{}")'.format(oneOrganizationURL, "Link"),
        'url': oneOrganizationURL,
        'category': category,
        # 'idealistURL': '=HYPERLINK("{}", "{}")'.format(idealistURL, "Link"),
        'idealistURL': idealistURL,
        'technologyList': toStr(technologyList),
        'topicsList': toStr(topicsList)
    })


threadPool = ThreadPoolExecutor(max_workers=1, thread_name_prefix="test_")

for mainPageOneOrganizationBS in tqdm(GSocOrganizationsBS.find_all('li')):
    threadPool.submit(infoGet, mainPageOneOrganizationBS)

    # break
threadPool.shutdown(wait=True)


wb = Workbook()
ws1 = wb.active
ws2 = wb.create_sheet(title='failed')
ws3 = wb.create_sheet(title='technologies')

ws1.title = 'succeed'

ws1.append(['组织名', 'GSoc页面链接', '分类', 'idea列表', '技术栈', '相关主题'])
for i, oneOrganization in enumerate(GSocOrganizationsInfoList):
    ws1.append([oneOrganization['title'],
                'Link',
                oneOrganization['category'],
                'Link',
                oneOrganization['technologyList'],
                oneOrganization['topicsList']
                ])
    ws1['B%d' % (i + 2)].hyperlink = oneOrganization['url']
    ws1['B%d' % (i + 2)].style = "Hyperlink"
    ws1['D%d' % (i + 2)].hyperlink = oneOrganization['idealistURL']
    ws1['D%d' % (i + 2)].style = "Hyperlink"

ws2.append(['组织名', 'GSoc页面链接'])
for oneOrganization in CrawlerErrorList:
    ws2.append([oneOrganization['title'], oneOrganization['url']])

ws3.append(['可用技术栈'])
for oneTechnology in sorted(AllTechnology):
    ws3.append([oneTechnology])

for ws in [ws1, ws2, ws3]:
    c = ws["A2"]
    ws.freeze_panes = c

wb.save('GSocOrganizations.xlsx')
